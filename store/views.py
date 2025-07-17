import datetime
import os
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.http import require_POST
from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required,user_passes_test
from django import forms
from django.db import transaction 
from django.conf import settings
from django.core.paginator import Paginator,EmptyPage, PageNotAnInteger
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.enums import TA_RIGHT, TA_LEFT, TA_CENTER
from decimal import Decimal

from .models import Brand, Category, Furniture,Order,OrderItem, Furniture, CustomUser, UserMembershipLog
from .forms import CategoryForm, CustomUserCreationForm, CustomUserUpdateForm, FurnitureForm, OrderForm


def category_list(request):
    top_level_categories = Category.objects.filter(parent__isnull=True).order_by('name')
    context = {
        'categories': top_level_categories,
    }
    return render(request, 'main/category_list.html', context)
from django.shortcuts import render, get_object_or_404
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
# Make sure to import your models correctly
from .models import Furniture, Category, Brand 

def furniture_list(request, category_slug=None, brand_slug=None):
    page_number = request.GET.get('page')
    request_category_id = request.GET.get('category_id')
    search_query = request.GET.get('q', '').strip()
    furniture_items = Furniture.objects.filter(is_visible=True)
    selected_category = None
    selected_brand = None

    # --- Category Filtering Logic ---
    if request_category_id:
        try:
            selected_category = Category.objects.get(pk=request_category_id)
            categories_to_include = [selected_category.pk]
            def get_all_children_pks(category_obj):
                children_pks = []
                for child in category_obj.children.all():
                    children_pks.append(child.pk)
                    children_pks.extend(get_all_children_pks(child))
                return children_pks
            categories_to_include.extend(get_all_children_pks(selected_category))
            furniture_items = furniture_items.filter(category__pk__in=categories_to_include)
        except Category.DoesNotExist:
            furniture_items = Furniture.objects.none()

    if brand_slug:
        try:
            selected_brand = get_object_or_404(Brand, slug=brand_slug)
            furniture_items = furniture_items.filter(brand=selected_brand)
        except Brand.DoesNotExist:
            furniture_items = Furniture.objects.none()

    # --- Search Filtering Logic ---
    if search_query:
        from django.db.models import Q
        furniture_items = furniture_items.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(item_code__icontains=search_query)
        )

    furniture_items = furniture_items.order_by('name')
    paginator = Paginator(furniture_items, 9)  # Show 9 items per page
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    categories = Category.objects.filter(parent__isnull=True).prefetch_related('children__children__children').order_by('name')
    recommended_items = Furniture.objects.filter(is_visible=True).order_by('?')[:8] # Get 8 random items
    featured_items = Furniture.objects.filter(is_visible=True, featured=True).order_by('-id')[:8] # Get 8 featured items
    own_brand_name = 'INHOUSE'

    try:
        own_brand = Brand.objects.get(name=own_brand_name)
        other_brands = Brand.objects.exclude(name=own_brand_name).order_by('name')
        all_brands_ordered = [own_brand] + list(other_brands)
    except Brand.DoesNotExist:
        all_brands_ordered = Brand.objects.all().order_by('name')
    context = {
        'page_obj': page_obj,
        'items': page_obj.object_list,
        'categories': categories,
        'brands': all_brands_ordered,
        'selected_category': selected_category,
        'selected_category_id': int(request_category_id) if request_category_id else None,
        'selected_brand': selected_brand,
        'recommended_items': recommended_items,
        'featured_items': featured_items,
        'search_query': search_query,
    }
    return render(request, 'main/furniture_list.html', context)

def furniture_detail(request,pk):
    item = get_object_or_404(Furniture,pk=pk)
    return render(request, 'main/furniture_detail.html',{'item':item})

def brand_detail(request, brand_slug):
    brand = get_object_or_404(Brand, slug=brand_slug)
    items = Furniture.objects.filter(brand=brand).order_by('name')

    categories = Category.objects.filter(parent__isnull=True).prefetch_related('children__children__children').order_by('name')
    all_brands = Brand.objects.all().order_by('name') # For the sidebar

    context = {
        'brand': brand,
        'items': items,
        'categories': categories, 
        'brands': all_brands, 
    }
    return render(request, 'main/brand_detail.html', context)

def add_to_cart(request, pk):
    cart = request.session.get('cart', {})
    pk_str = str(pk)

    try:
        quantity = int(request.POST.get('quantity', 1))
    except ValueError:
        quantity = 1

    if pk_str in cart:
        cart[pk_str] += quantity
    else:
        cart[pk_str] = quantity

    request.session['cart'] = cart
    print(pk_str, "your cart now has", cart[pk_str], "items.")
    return redirect('furniture_list')

def view_cart(request):
    cart = request.session.get('cart',{})
    items=[]
    total=0
    for pk, qty in cart.items():
        try:
            furniture = Furniture.objects.get(pk=pk)
            subtotal = furniture.price * qty
            total += subtotal
            items.append({
                'furniture': furniture,
                'quantity': qty,
                'subtotal': subtotal
            })
        except Furniture.DoesNotExist:
            continue
    return render(request, 'main/cart.html',{'items':items,'total':total})

@require_POST
def update_cart(request,pk):
    cart= request.session.get('cart',{})
    quantity = int(request.POST.get('quantity',1))
    if quantity >0:
        cart[str(pk)]=quantity
    else:
        cart.pop(str(pk),None)
    request.session['cart']=cart
    return redirect('view_cart')

def remove_from_cart(request,pk):
    cart = request.session.get('cart',{})
    cart.pop(str(pk),None)
    request.session['cart']= cart
    return redirect('view_cart')

from django.contrib.auth import login
from django.core.mail import send_mail

def register(request):
    next_url = request.GET.get('next') or request.POST.get('next') or 'furniture_list'
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            # Send welcome email
            send_mail(
                'Welcome to Our Furniture Store!',
                f'Hi {user.username}, thanks for registering!',
                'kgmobile.service@gmial.com',
                [user.email],
                fail_silently=True
            )
            return redirect(next_url)
    else:
        form = CustomUserCreationForm()
    return render(request, 'main/register.html', {'form': form})

from django.contrib.auth.decorators import login_required

@login_required
def profile(request):
    member_level_colors = {
        'Gold': 'bg-yellow-100 border-yellow-300',    # Light yellow background for Gold
        'Silver': 'bg-gray-100 border-gray-300',     # Light gray for Silver
        'Bronze': 'bg-amber-50 border-amber-200',    # Very light orange/amber for Bronze
        'Platinum': 'bg-blue-100 border-blue-300',   # Light blue for Platinum
        'Default': 'bg-white border-gray-200',      # Default white background
    }
    user_member_level = request.user.member_level if request.user.member_level else 'Default'
    background_classes = member_level_colors.get(user_member_level, member_level_colors['Default'])
    context = {
        'background_classes': background_classes,
    }
    return render(request, 'main/profile.html', context)

def get_session_cart(request):
    return request.session.get('cart', {}) # cart = {'furniture_id': quantity, ...}
def clear_session_cart(request):
    if 'cart' in request.session:
        del request.session['cart']

@login_required
def checkout_view(request):
    cart = get_session_cart(request)
    if not cart:
        return redirect('furniture_list') # Or some other appropriate page
    cart_items_data = []
    total_cart_amount = 0
    for furniture_id, quantity in cart.items():
        try:
            furniture = Furniture.objects.get(id=furniture_id)
            item_total = furniture.price * quantity
            cart_items_data.append({
                'furniture': furniture,
                'quantity': quantity,
                'price': furniture.price,
                'total': item_total
            })
            total_cart_amount += item_total
        except Furniture.DoesNotExist:
            continue

    if request.method == 'POST':
        form = OrderForm(request.POST)
        if form.is_valid():
            with transaction.atomic():
                order = Order.objects.create(
                    user=request.user,
                    customer_name=form.cleaned_data['customer_name'],
                    customer_address=form.cleaned_data['customer_address'],
                    customer_phone=form.cleaned_data['customer_phone'],
                    payment_method=form.cleaned_data['payment_method'],
                    total_amount=total_cart_amount, # Set calculated total
                    status='Pending',
                    payment_status='Pending'
                )
                for item_data in cart_items_data:
                    OrderItem.objects.create(
                        order=order,
                        furniture=item_data['furniture'],
                        quantity=item_data['quantity'],
                        price_at_purchase=item_data['price'] # Use price at time of purchase
                    )
                clear_session_cart(request) # Clear cart after successful order
                return redirect('main/order_success', order_id=order.id) # Redirect to success page
    else:
        initial_data = {
            'customer_name': request.user.username,
            'customer_address': request.user.address if hasattr(request.user, 'address') else '', # Assuming user might have address
            'customer_phone': request.user.phone,
        }
        form = OrderForm(initial=initial_data)
    context = {
        'form': form,
        'cart_items_data': cart_items_data,
        'total_cart_amount': total_cart_amount,
    }
    return render(request, 'main/checkout.html', context)

@login_required
def order_success_view(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    return render(request, 'main/order_success.html', {'order': order})

@login_required
def user_order_history(request):
    orders = Order.objects.filter(user=request.user).order_by('-created_at')
    context = {'orders': orders}
    return render(request, 'main/order_history.html', context)

# before order cancellation feature added
@login_required
def order_detail(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    context = {'order': order}
    return render(request, 'main/order_detail.html', context)

def generate_invoice_pdf(request, order_id):
    order = get_object_or_404(Order, id=order_id, user=request.user)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Invoice_Order_{order.order_number}.pdf"'

    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4
    styles = getSampleStyleSheet()
    
    GOLD_COLOR = colors.HexColor("#DAA520") # Primary accent gold
    DARK_GRAY = colors.HexColor("#333333")  # Dark text/header elements
    LIGHT_GRAY = colors.HexColor("#F8F8F8") # Light background for alternating rows
    BORDER_GRAY = colors.HexColor("#E0E0E0") # Light border color
    MEDIUM_GRAY = colors.HexColor("#666666") # Secondary text color
    WATERMARK_COLOR = colors.HexColor("#DAA520") # Use gold for watermark, or a very light gray
    logo_path = os.path.join(settings.STATIC_ROOT, 'img', 'logop.png') # Adjust this path

    # --- Company Information ---
    company_name = "INHOUSE Luxury Furniture"
    company_address_line1 = "No.150/151, Bo Aung Kyaw Road,Shwe Than Lwin Industrial Zone,"
    company_address_line2 = "Hlaing Thar Yar Tsp, Yangon, Myanamr."
    company_phone = "+959-443887880,+959-443887881"
    company_email = "info@inhousemm.com"

    p.saveState() # Save the current canvas state before applying transformations
    p.setFillColor(WATERMARK_COLOR)
    p.setStrokeColor(WATERMARK_COLOR)
    p.setDash(1, 0) # Solid line (no dashes)

    p.setFillAlpha(0.1) # Slightly lower opacity for a full-page watermark
    p.setStrokeAlpha(0.1)

    p.translate(width/2.0, height/2.0)
    p.rotate(45) # Rotate 45 degrees

    watermark_text = "INHOUSE"
    font_size = 30
    p.setFont("Helvetica-Bold", font_size)


    text_width_approx = p.stringWidth(watermark_text, "Helvetica-Bold", font_size)
    text_height_approx = font_size * 1.2

    x_spacing = text_width_approx * 1.5
    y_spacing = text_height_approx * 3.0

    
    # Loop over X-axis
    for x_offset in range(int(-width * 0.7), int(width * 0.7), int(x_spacing)):
        # Loop over Y-axis
        for y_offset in range(int(-height * 0.7), int(height * 0.7), int(y_spacing)):
            p.drawCentredString(x_offset, y_offset, watermark_text)

    p.restoreState() 
    # === End Watermark ===

########################logo image
    try:
            if os.path.exists(logo_path):
                logo_width = 6 * cm  
                logo_height = 2.5 * cm 
                p.drawImage(logo_path,width- 8 * cm, height - 6.5 * cm, width=logo_width, height=logo_height, preserveAspectRatio=True)
    
            else:
                print(f"Logo file not found at: {logo_path}")
                company_info_y_offset = 3.5 * cm
                invoice_title_y_offset = 3.5 * cm

    except Exception as e:
        print(f"Error drawing logo: {e}") # For debugging
        company_info_y_offset = 3.5 * cm
        invoice_title_y_offset = 3.5 * cm

    # === Header & Company Details ===
    p.setFont("Helvetica-Bold", 32) # Larger, more impactful invoice title
    p.setFillColor(GOLD_COLOR)
    p.drawRightString(width - 2.5 * cm, height - 3.5 * cm, "INVOICE") # Right-aligned

    p.setFont("Helvetica-Bold", 14)
    p.setFillColor(DARK_GRAY)
    p.drawString(2.5 * cm, height - 5.5 * cm, company_name)
    p.setFont("Helvetica", 9) 
    p.setFillColor(MEDIUM_GRAY)
    p.drawString(2.5 * cm, height - 6.0 * cm, company_address_line1)
    p.drawString(2.5 * cm, height - 6.4 * cm, company_address_line2)
    p.drawString(2.5 * cm, height - 6.8 * cm, f"Phone: {company_phone}")
    p.drawString(2.5 * cm, height - 7.2 * cm, f"Email: {company_email}")

    # Horizontal line separator below company details
    p.setStrokeColor(BORDER_GRAY)
    p.setLineWidth(0.7) # Slightly thicker line
    p.line(2.5 * cm, height - 7.8 * cm, width - 2.5 * cm, height - 7.8 * cm)

    # === Invoice Details & Customer Info ===
    y_start_info_block = height - 8.8 * cm

    # Invoice Details (Right side, aligned with invoice title)
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(DARK_GRAY)
    p.drawRightString(width - 2.5 * cm, y_start_info_block, "Invoice Details:")
    p.setFont("Helvetica", 10)
    p.setFillColor(MEDIUM_GRAY)
    p.drawRightString(width - 2.5 * cm, y_start_info_block - 0.6 * cm, f"Invoice No: {order.order_number}")
    p.drawRightString(width - 2.5 * cm, y_start_info_block - 1.2 * cm, f"Date: {order.created_at.strftime('%B %d, %Y')}")
    p.drawRightString(width - 2.5 * cm, y_start_info_block - 1.8 * cm, f"Time: {order.created_at.strftime('%H:%M')}")
    p.drawRightString(width - 2.5 * cm, y_start_info_block - 2.4 * cm, f"Status: {order.get_status_display()}")
    p.drawRightString(width - 2.5 * cm, y_start_info_block - 3.0 * cm, f"Payment Status: {order.get_payment_status_display()}")


    # Customer Info (Left side)
    p.setFont("Helvetica-Bold", 11)
    p.setFillColor(DARK_GRAY)
    p.drawString(2.5 * cm, y_start_info_block, "Bill To:")
    p.setFont("Helvetica", 10)
    p.setFillColor(MEDIUM_GRAY)
    p.drawString(2.5 * cm, y_start_info_block - 0.6 * cm, order.customer_name)
    p.drawString(2.5 * cm, y_start_info_block - 1.2 * cm, order.customer_address)
    p.drawString(2.5 * cm, y_start_info_block - 1.8 * cm, f"Phone: {order.customer_phone}")
    p.drawString(2.5 * cm, y_start_info_block - 2.4 * cm, f"Payment Method: {order.payment_method}")

    # --- Order Items Table ---
    data = [
        [Paragraph("<b>Item Description</b>", styles['Normal']), Paragraph("<b>Qty</b>", styles['Normal']), Paragraph("<b>Unit Price</b>", styles['Normal']), Paragraph("<b>Total</b>", styles['Normal'])]
    ]
    for item in order.items.all():
        data.append([
            Paragraph(item.furniture.name, styles['Normal']),
            str(item.quantity),
            f"${item.price_at_purchase:.2f}",
            f"${item.get_total():.2f}"
        ])

    table_y_start = y_start_info_block - 4.5 * cm # Position table below info blocks
    
    # Define table style
    table_style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GOLD_COLOR), # Header background
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white), # Header text color
        ('ALIGN', (0, 0), (0, -1), 'LEFT'), # Item Description left-aligned
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'), # Qty, Unit Price, Total centered
        ('ALIGN', (2, 0), (-1, -1), 'RIGHT'), # Unit Price and Total right-aligned for numbers
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'), # Header font
        ('FONTSIZE', (0, 0), (-1, 0), 10), # Header font size
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8), # Header padding
        ('TOPPADDING', (0, 0), (-1, 0), 8), # Header padding
        ('GRID', (0, 0), (-1, -1), 0.5, BORDER_GRAY), # All borders
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), # Vertical alignment
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'), # Body font
        ('FONTSIZE', (0, 1), (-1, -1), 9), # Body font size
    ])

    # Apply alternating row background
    for i in range(1, len(data)):
        if i % 2 == 0:
            table_style.add('BACKGROUND', (0, i), (-1, i), LIGHT_GRAY)
        else:
            table_style.add('BACKGROUND', (0, i), (-1, i), colors.white)


    # Column widths (adjust as needed for content)
    col_widths = [8.5 * cm, 2 * cm, 3 * cm, 3 * cm] # Item, Qty, Unit Price, Total

    table = Table(data, colWidths=col_widths)
    table.setStyle(table_style)

    # Draw table
    table_width, table_height = table.wrapOn(p, width, height)
    p.saveState() # Save current state
    table.drawOn(p, 2.5 * cm, table_y_start - table_height) # Position table
    p.restoreState() # Restore state

    current_y = table_y_start - table_height

    # --- Total Section ---
    total_section_start_y = current_y - 1.5 * cm

    subtotal = sum(item.get_total() for item in order.items.all())
    tax_rate = Decimal('0.05') 
    shipping_cost = Decimal('25.00') 
    tax_amount = subtotal * tax_rate
    grand_total = subtotal + tax_amount + shipping_cost

    # Draw lines for totals
    p.setStrokeColor(BORDER_GRAY)
    p.setLineWidth(0.5)
    
    # Subtotal, Tax, Shipping
    p.setFont("Helvetica", 10)
    p.setFillColor(DARK_GRAY)
    p.drawRightString(width - 5 * cm, total_section_start_y, "Subtotal:")
    p.drawRightString(width - 2.5 * cm, total_section_start_y, f"${subtotal:.2f}")

    p.drawRightString(width - 5 * cm, total_section_start_y - 0.7 * cm, f"Tax ({tax_rate * 100:.0f}%):")
    p.drawRightString(width - 2.5 * cm, total_section_start_y - 0.7 * cm, f"${tax_amount:.2f}")

    p.drawRightString(width - 5 * cm, total_section_start_y - 1.4 * cm, "Shipping:")
    p.drawRightString(width - 2.5 * cm, total_section_start_y - 1.4 * cm, f"${shipping_cost:.2f}")

    # Line above Grand Total
    p.setLineWidth(1) # Thicker line
    p.line(width - 7.5 * cm, total_section_start_y - 1.8 * cm, width - 2.5 * cm, total_section_start_y - 1.8 * cm)

    # Grand Total
    p.setFont("Helvetica-Bold", 16) # Larger, bolder for grand total
    p.setFillColor(GOLD_COLOR)
    p.drawRightString(width - 5 * cm, total_section_start_y - 2.8 * cm, "GRAND TOTAL:")
    p.drawRightString(width - 2.5 * cm, total_section_start_y - 2.8 * cm, f"${grand_total:.2f}")


    # === Footer ===
    p.setFillColor(MEDIUM_GRAY)
    p.setFont("Helvetica-Oblique", 9)
    p.setLineWidth(0.5)
    p.line(2.5 * cm, 2.5 * cm, width - 2.5 * cm, 2.5 * cm) # Line above footer
    p.drawCentredString(width / 2, 1.5 * cm, "Thank you for choosing INHOUSE Luxury Furniture. We appreciate your business!")

    p.showPage()
    p.save()
    return response

################################staff view session #######################################
def is_staff_check(user):
    return user.is_authenticated and user.is_staff
@login_required
@user_passes_test(is_staff_check)
def staff_dashboard(request):
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='Pending').count()
    delivered_orders_today = Order.objects.filter(
        status='Delivered', created_at__date=datetime.date.today()
    ).count()
    total_customers = CustomUser.objects.count()
    total_furniture_items = Furniture.objects.count()

    recent_orders = Order.objects.order_by('-created_at')[:10]
    recent_membership_changes = UserMembershipLog.objects.order_by('-changed_at')[:5]

    total_categories = Category.objects.count()
    recent_category_logs = []
    context = {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'delivered_orders_today': delivered_orders_today,
        'total_customers': total_customers,
        'total_furniture_items': total_furniture_items,
        'recent_orders': recent_orders,
        'recent_membership_changes': recent_membership_changes,
        'total_categories': total_categories,
        'recent_category_logs': recent_category_logs,
    }
    return render(request, 'staff/dashboard.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_order_list(request):
    """
    Staff view to list all orders with filtering options.
    """
    orders = Order.objects.all().select_related('user').order_by('-created_at')

    status_filter = request.GET.get('status')
    payment_status_filter = request.GET.get('payment_status')
    search_query = request.GET.get('q')

    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query) |
            Q(user__username__icontains=search_query) # Search by associated username
        )

    context = {
        'orders': orders,
        'status_choices': Order.STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_CHOICES,
        'selected_status': status_filter,
        'selected_payment_status': payment_status_filter,
        'search_query': search_query,
    }
    return render(request, 'staff/order_list.html', context)

from django.db.models import Q

@login_required
@user_passes_test(is_staff_check)
def staff_order_detail(request, order_id):
    """
    Staff view to see order details and update status/payment status.
    """
    order = get_object_or_404(Order.objects.select_related('user'), id=order_id)
    order_items = order.items.select_related('furniture') # Get related order items

    

    if request.method == 'POST':
        form = OrderForm(request.POST) # Reusing OrderForm, might need a specific staff form
        if 'update_status' in request.POST:
            new_status = request.POST.get('status')
            new_payment_status = request.POST.get('payment_status')

            # Validate choices against model fields
            if new_status in [choice[0] for choice in Order.STATUS_CHOICES] and \
               new_payment_status in [choice[0] for choice in Order.PAYMENT_CHOICES]:

                old_status = order.status
                old_payment_status = order.payment_status
                order.status = new_status
                order.payment_status = new_payment_status
                order.save(update_fields=['status', 'payment_status', 'updated_at'])


                messages.success(request, f"Order #{order.order_number} status updated successfully.")
                return redirect('staff_order_detail', order_id=order.id)
            else:
                messages.error(request, "Invalid status or payment status selected.")

        elif 'cancel_order' in request.POST:
            if order.status not in ['Delivered', 'Cancelled']: # Prevent cancelling already delivered/cancelled orders
                order.status = 'Cancelled'
                order.payment_status = 'Failed' # Or 'Refunded', depending on your flow
                order.save(update_fields=['status', 'payment_status', 'updated_at'])
                
                messages.warning(request, f"Order #{order.order_number} has been cancelled.")
            else:
                messages.error(request, "Cannot cancel an order that is already Delivered or Cancelled.")
            return redirect('staff_order_detail', order_id=order.id)

    context = {
        'order': order,
        'order_items': order_items,
        'status_choices': Order.STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_CHOICES,
    }
    return render(request, 'staff/order_detail.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_furniture_list(request):
    furniture_items = Furniture.objects.all().select_related('category', 'brand').order_by('name')

    search_query = request.GET.get('q')
    category_id = request.GET.get('category')
    brand_id = request.GET.get('brand')
    featured_filter = request.GET.get('featured')
    visible_filter = request.GET.get('is_visible')

    if search_query:
        furniture_items = furniture_items.filter(
            Q(name__icontains=search_query) |
            Q(item_code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    if category_id:
        furniture_items = furniture_items.filter(category__id=category_id)
    if brand_id:
        furniture_items = furniture_items.filter(brand__id=brand_id)
    if featured_filter == 'yes':
        furniture_items = furniture_items.filter(featured=True)
    elif featured_filter == 'no':
        furniture_items = furniture_items.filter(featured=False)
    if visible_filter == 'on':
        furniture_items = furniture_items.filter(is_visible=True)
    elif visible_filter == 'off':
        furniture_items = furniture_items.filter(is_visible=False)


    categories = Category.objects.all().order_by('name')
    brands = Brand.objects.all().order_by('name')
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    # --- Export Excel/PDF ---
    export_format = request.GET.get('export')
    if export_format in ['excel', 'pdf']:
        import io
        from django.http import FileResponse
        if export_format == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Furniture List'

            # Title row
            title = 'Furniture List Export'
            ws.merge_cells('A1:H1')
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill('solid', fgColor='2d3748')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Timestamp row
            import datetime
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.merge_cells('A2:H2')
            ws['A2'] = f'Exported: {timestamp}'
            ws['A2'].font = Font(size=11, italic=True, color='4a5568')
            ws['A2'].fill = PatternFill('solid', fgColor='f6e05e')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

            # Header row with row numbering
            headers = ['No.', 'Name', 'Item Code', 'Category', 'Brand', 'Featured', 'Status', 'Price']
            ws.append(headers)
            header_row = 3
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = Font(size=14, bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='f6ad55')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data rows with row numbering
            for idx, item in enumerate(furniture_items, start=1):
                ws.append([
                    idx,
                    item.name,
                    item.item_code,
                    item.category.name if item.category else '',
                    item.brand.name if item.brand else '',
                    'Yes' if item.featured else 'No',
                    'Yes' if item.is_visible else 'No',
                    str(item.price)
                ])

            # Table styling: borders and alignment
            beautiful_border = Border(
                left=Side(style='medium', color='f6ad55'),
                right=Side(style='medium', color='f6ad55'),
                top=Side(style='medium', color='f6ad55'),
                bottom=Side(style='medium', color='f6ad55'),
                
            )
            max_row = ws.max_row
            max_col = ws.max_column
            for row in ws.iter_rows(min_row=header_row, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = beautiful_border
                    # No. column center, Price column right, others left
                    if cell.column == 1:
                        cell.alignment = Alignment(vertical='center', horizontal='center')
                    elif cell.column == 8:
                        cell.alignment = Alignment(vertical='center', horizontal='right')
                    else:
                        cell.alignment = Alignment(vertical='center', horizontal='left')

            # Set column widths for better fit, name and featured columns fit to text
            name_col_idx = 2
            featured_col_idx = 6
            max_name_len = max(len(str(ws.cell(row=row, column=name_col_idx).value)) for row in range(header_row+1, ws.max_row+1)) if ws.max_row > header_row else 10
            name_col_width = min(max(15, max_name_len + 2), 40)  # min 15, max 40
            # Set featured column width to match header text length
            featured_header = headers[featured_col_idx - 1]
            featured_col_width = max(8, len(featured_header) + 4)  # min 8, header length + 2
            col_widths = [7, name_col_width, 15, 18, 18, featured_col_width, 10, 12]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"furniture_list_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = FileResponse(output, as_attachment=True, filename=filename)
            return response
        elif export_format == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            from reportlab.platypus import Table, TableStyle, Paragraph
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            import datetime
            output = io.BytesIO()
            page_size = landscape(A4)
            p = canvas.Canvas(output, pagesize=page_size)
            width, height = page_size
            styles = getSampleStyleSheet()
            # Title
            p.setFont('Helvetica-Bold', 18)
            p.setFillColor(colors.HexColor('#2d3748'))
            p.drawString(2 * cm, height - 2 * cm, 'Furniture List Export')
            # Timestamp
            p.setFont('Helvetica', 10)
            p.setFillColor(colors.HexColor('#4a5568'))
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            p.drawString(2 * cm, height - 2.7 * cm, f'Exported: {timestamp}')
            # Table Data
            headers = ['No.', 'Name', 'Item Code', 'Category', 'Brand', 'Featured', 'Visible', 'Price']
            data = [headers]
            for idx, item in enumerate(furniture_items, start=1):
                data.append([
                    str(idx),
                    item.name,
                    item.item_code,
                    item.category.name if item.category else '',
                    item.brand.name if item.brand else '',
                    'Yes' if item.featured else 'No',
                    'Yes' if item.is_visible else 'No',
                    str(item.price)
                ])
            # Table Style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f6ad55')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),      # Name left
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),      # Item Code left
                ('ALIGN', (2, 0), (5, -1), 'CENTER'),    # Category, Brand, Featured, Visible center
                ('ALIGN', (6, 0), (6, -1), 'RIGHT'),     # Price right
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            # Dynamically set name column width based on max text length (tighter fit)
            max_name_len = max([len(str(row[1])) for row in data]) if data else 10
            name_col_width = min(max(2*cm, max_name_len*0.18*cm), 7*cm)  # Reduce minimum and maximum, and scale
            # Dynamically set featured column width
            max_featured_len = max([len(str(row[5])) for row in data]) if data else 3
            featured_col_width = min(max(1.2*cm, max_featured_len*0.35*cm), 2.5*cm)  # min 1.2cm, max 2.5cm
            col_widths = [1.5*cm, name_col_width, 3*cm, 4*cm, 4*cm, featured_col_width, 2*cm, 3*cm]

            # Estimate rows per page (landscape A4)
            # Further reduce margins to fit more rows
            row_height = 0.85 * cm  # Slightly smaller per row
            top_margin = 2.5 * cm   # Reduce top margin
            bottom_margin = 0.9 * cm  # Reduce bottom margin for footer text
            line_height = 0.6 * cm     # Reduce space for horizontal line above footer
            available_height = height - (top_margin + bottom_margin + line_height)
            rows_per_page = int(available_height // row_height) + 3
            # Always include header row on each page
            data_rows = data[1:]
            total_rows = len(data_rows)
            pages = [data_rows[i:i+rows_per_page-1] for i in range(0, total_rows, rows_per_page-1)]

            for page_num, page_rows in enumerate(pages):
                # Table data for this page: header + chunk
                page_data = [headers] + page_rows
                # Table style: alternating row colors
                page_table_style = TableStyle(table_style.getCommands())
                for i in range(1, len(page_data)):
                    bg_color = colors.HexColor('#f7fafc') if i % 2 == 0 else colors.white
                    page_table_style.add('BACKGROUND', (0, i), (-1, i), bg_color)
                table = Table(page_data, colWidths=col_widths)
                table.setStyle(page_table_style)
                # Draw table
                table_width, table_height = table.wrapOn(p, width, height)
                p.saveState()
                table.drawOn(p, 2 * cm, height - 4 * cm - table_height)
                p.restoreState()
                # Footer: horizontal line, text and page number on same row
                p.setLineWidth(0.5)
                p.setStrokeColor(colors.HexColor('#e2e8f0'))
                line_y = 1.7 * cm  # Slightly above text
                p.line(2 * cm, line_y, width - 2 * cm, line_y)
                # Footer text center, page number right on same row
                footer_y = 1.3 * cm
                p.setFont('Helvetica-Oblique', 9)
                p.setFillColor(colors.HexColor('#718096'))
                footer_text = "Generated by INHOUSE Portal"
                p.drawCentredString(width / 2, footer_y, footer_text)
                # Page number right-aligned on same row
                p.setFont('Helvetica', 8)
                p.setFillColor(colors.HexColor('#4a5568'))
                page_num_text = f"Page {page_num + 1} of {len(pages)}"
                p.drawRightString(width - 2 * cm, footer_y, page_num_text)
                p.showPage()
            p.save()
            output.seek(0)
            timestamp_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"furniture_list_{timestamp_str}.pdf"
            response = FileResponse(output, as_attachment=True, filename=filename)
            return response

    paginator = Paginator(furniture_items, 10) # Show 10 items per page
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj, 
        'items': page_obj.object_list, 
        'categories': categories,
        'brands': brands,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_brand': brand_id,
        'selected_featured': featured_filter,
        'selected_is_visible':visible_filter,
    }
    return render(request, 'staff/furniture_list.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_furniture_manage(request, pk=None):
    if pk:
        furniture = get_object_or_404(Furniture, pk=pk)
        action = "Edit"
    else:
        furniture = None
        action = "Add"

    if request.method == 'POST':
        form = FurnitureForm(request.POST, request.FILES, instance=furniture)
        if form.is_valid():
            furniture_item = form.save()
            messages.success(request, f"Furniture item '{furniture_item.name}' saved successfully.")
            return redirect('staff_furniture_list')
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = FurnitureForm(instance=furniture)

    context = {
        'form': form,
        'action': action,
        'furniture': furniture, 
    }
    return render(request, 'staff/furniture_form.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_furniture_delete(request, pk):
    furniture = get_object_or_404(Furniture, pk=pk)
    if request.method == 'POST':
        furniture.delete()
        messages.success(request, f"Furniture item '{furniture.name}' deleted successfully.")
        return redirect('staff_furniture_list')
    context = {
        'furniture': furniture
    }
    return render(request, 'staff/furniture_confirm_delete.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_customer_list(request):
    # customers = CustomUser.objects.filter(is_staff=False).order_by('username')
    customers = CustomUser.objects.filter().order_by('username')

    search_query = request.GET.get('q', '').strip()
    selected_member_level = request.GET.get('member_level', '')

    filters = []

    if search_query:
        filters.append(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) | 
            Q(c_code__icontains=search_query) 
        )

    # Apply Member Level Filter
    if selected_member_level:
        filters.append(Q(member_level=selected_member_level)) 

    # Apply all collected filters
    if filters:
        customers = customers.filter(*filters).distinct() # Use distinct() to avoid duplicates

    # Pagination
    paginator = Paginator(customers, 15) # Show 15 customers per page
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'customers': page_obj.object_list, # The customers for the current page
        'search_query': search_query,
        'selected_member_level': selected_member_level,
        'member_level_choices': CustomUser.member_level_choices, # Get choices directly from CustomUser model
    }
    return render(request, 'staff/customer_list.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_customer_detail(request, pk):
    customer = get_object_or_404(CustomUser, pk=pk)
    customer_orders = customer.orders.all().order_by('-created_at')[:10]

    if request.method == 'POST':
        form = CustomUserUpdateForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f"Customer '{customer.username}' updated successfully.")
            return redirect('staff_customer_detail', pk=customer.pk)
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = CustomUserUpdateForm(instance=customer)

    context = {
        'customer': customer,
        'form': form,
        'customer_orders': customer_orders,
    }
    return render(request, 'staff/customer_detail.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_category_list(request):
    all_categories = Category.objects.select_related('parent').order_by('parent__name', 'name')
    top_level_categories = [cat for cat in all_categories if cat.parent is None]
    parent_groups = {}
    for cat in all_categories:
        if cat.parent:
            parent_groups.setdefault(cat.parent, []).append(cat)

    form = CategoryForm()
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            return redirect('staff_category_list')

    context = {
        'top_level_categories': top_level_categories,
        'parent_groups': parent_groups,
        'form': form,
    }
    return render(request, 'staff/category_list.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            # log_category_save(sender=Category, instance=category, created=False, user=request.user)
            return redirect('staff_category_list')
    else:
        form = CategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
    }
    return render(request, 'staff/category_edit.html', context) # Create this template

@login_required
@user_passes_test(is_staff_check)
def staff_category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category_name_for_log = category.name # Store name before deletion
        category.delete()
      
        return redirect('staff_category_list')
    context = {
        'category': category,
    }
    return render(request, 'staff/category_confirm_delete.html', context) # Create this template

