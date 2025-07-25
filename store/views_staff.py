import json
import subprocess,os,datetime
import tempfile

from django.shortcuts import get_object_or_404, render,redirect
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import FileResponse
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db.models import Q
from django.contrib.admin.views.decorators import staff_member_required

from .models import Brand, Category, ExchangeRate, Furniture,Order,OrderItem, Furniture, CustomUser, UserMembershipLog
from .forms import CategoryForm, CustomUserCreationForm, CustomUserUpdateForm, FurnitureForm, OrderForm

def is_staff_check(user):
    return user.is_authenticated and user.is_staff
@login_required
@user_passes_test(is_staff_check)
def staff_dashboard(request):
    total_orders = Order.objects.count()
    pending_orders = Order.objects.filter(status='Pending').count()
    delivered_orders_today = Order.objects.filter(
        status='Delivered', created_at__date=datetime.date.today()
    ).count()
    total_customers = CustomUser.objects.count()
    total_furniture_items = Furniture.objects.count()

    recent_orders = Order.objects.order_by('-created_at')[:10]
    recent_membership_changes = UserMembershipLog.objects.order_by('-changed_at')[:5]

    total_categories = Category.objects.count()
    recent_category_logs = []
    context = {
        'total_orders': total_orders,
        'pending_orders': pending_orders,
        'delivered_orders_today': delivered_orders_today,
        'total_customers': total_customers,
        'total_furniture_items': total_furniture_items,
        'recent_orders': recent_orders,
        'recent_membership_changes': recent_membership_changes,
        'total_categories': total_categories,
        'recent_category_logs': recent_category_logs,
    }
    return render(request, 'staff/dashboard.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_order_list(request):
    """
    Staff view to list all orders with filtering options.
    """
    orders = Order.objects.all().select_related('user').order_by('-created_at')

    status_filter = request.GET.get('status')
    payment_status_filter = request.GET.get('payment_status')
    search_query = request.GET.get('q')

    if status_filter:
        orders = orders.filter(status=status_filter)
    if payment_status_filter:
        orders = orders.filter(payment_status=payment_status_filter)
    if search_query:
        orders = orders.filter(
            Q(order_number__icontains=search_query) |
            Q(customer_name__icontains=search_query) |
            Q(customer_phone__icontains=search_query) |
            Q(user__username__icontains=search_query) # Search by associated username
        )

    context = {
        'orders': orders,
        'status_choices': Order.STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_CHOICES,
        'selected_status': status_filter,
        'selected_payment_status': payment_status_filter,
        'search_query': search_query,
    }
    return render(request, 'staff/order_list.html', context)

from django.db.models import Q

@login_required
@user_passes_test(is_staff_check)
def staff_order_detail(request, order_id):

    order = get_object_or_404(Order.objects.select_related('user'), id=order_id)
    order_items = order.items.select_related('furniture') # Get related order items

    

    if request.method == 'POST':
        form = OrderForm(request.POST) # Reusing OrderForm, might need a specific staff form
        if 'update_status' in request.POST:
            new_status = request.POST.get('status')
            new_payment_status = request.POST.get('payment_status')

            if new_status in [choice[0] for choice in Order.STATUS_CHOICES] and \
               new_payment_status in [choice[0] for choice in Order.PAYMENT_CHOICES]:

                old_status = order.status
                old_payment_status = order.payment_status
                order.status = new_status
                order.payment_status = new_payment_status
                order.save(update_fields=['status', 'payment_status', 'updated_at'])


                messages.success(request, f"Order #{order.order_number} status updated successfully.")
                return redirect('staff_order_detail', order_id=order.id)
            else:
                messages.error(request, "Invalid status or payment status selected.")

        elif 'cancel_order' in request.POST:
            if order.status not in ['Delivered', 'Cancelled']: # Prevent cancelling already delivered/cancelled orders
                order.status = 'Cancelled'
                order.payment_status = 'Failed' # Or 'Refunded', depending on your flow
                order.save(update_fields=['status', 'payment_status', 'updated_at'])
                
                messages.warning(request, f"Order #{order.order_number} has been cancelled.")
            else:
                messages.error(request, "Cannot cancel an order that is already Delivered or Cancelled.")
            return redirect('staff_order_detail', order_id=order.id)

    context = {
        'order': order,
        'order_items': order_items,
        'status_choices': Order.STATUS_CHOICES,
        'payment_status_choices': Order.PAYMENT_CHOICES,
    }
    return render(request, 'staff/order_detail.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_furniture_list(request):
    furniture_items = Furniture.objects.all().select_related('category', 'brand').order_by('name')

    search_query = request.GET.get('q')
    category_id = request.GET.get('category')
    brand_id = request.GET.get('brand')
    featured_filter = request.GET.get('featured')
    visible_filter = request.GET.get('is_visible')

    if search_query:
        furniture_items = furniture_items.filter(
            Q(name__icontains=search_query) |
            Q(item_code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    if category_id:
        furniture_items = furniture_items.filter(category__id=category_id)
    if brand_id:
        furniture_items = furniture_items.filter(brand__id=brand_id)
    if featured_filter == 'yes':
        furniture_items = furniture_items.filter(featured=True)
    elif featured_filter == 'no':
        furniture_items = furniture_items.filter(featured=False)
    if visible_filter == 'on':
        furniture_items = furniture_items.filter(is_visible=True)
    elif visible_filter == 'off':
        furniture_items = furniture_items.filter(is_visible=False)


    categories = Category.objects.all().order_by('name')
    brands = Brand.objects.all().order_by('name')
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger

    # --- Export Excel/PDF ---
    export_format = request.GET.get('export')
    if export_format in ['excel', 'pdf']:
        import io
        from django.http import FileResponse
        if export_format == 'excel':
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Furniture List'

            # Title row
            title = 'Furniture List Export'
            ws.merge_cells('A1:H1')
            ws['A1'] = title
            ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
            ws['A1'].fill = PatternFill('solid', fgColor='2d3748')
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

            # Timestamp row
            import datetime
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            ws.merge_cells('A2:H2')
            ws['A2'] = f'Exported: {timestamp}'
            ws['A2'].font = Font(size=11, italic=True, color='4a5568')
            ws['A2'].fill = PatternFill('solid', fgColor='f6e05e')
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

            # Header row with row numbering
            headers = ['No.', 'Name', 'Item Code', 'Category', 'Brand', 'Featured', 'Status', 'Price']
            ws.append(headers)
            header_row = 3
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=header_row, column=col_num)
                cell.font = Font(size=14, bold=True, color='FFFFFF')
                cell.fill = PatternFill('solid', fgColor='f6ad55')
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Data rows with row numbering
            for idx, item in enumerate(furniture_items, start=1):
                ws.append([
                    idx,
                    item.name,
                    item.item_code,
                    item.category.name if item.category else '',
                    item.brand.name if item.brand else '',
                    'Yes' if item.featured else 'No',
                    'Yes' if item.is_visible else 'No',
                    str(item.price)
                ])

            # Table styling: borders and alignment
            beautiful_border = Border(
                left=Side(style='medium', color='f6ad55'),
                right=Side(style='medium', color='f6ad55'),
                top=Side(style='medium', color='f6ad55'),
                bottom=Side(style='medium', color='f6ad55'),
                
            )
            max_row = ws.max_row
            max_col = ws.max_column
            for row in ws.iter_rows(min_row=header_row, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = beautiful_border
                    # No. column center, Price column right, others left
                    if cell.column == 1:
                        cell.alignment = Alignment(vertical='center', horizontal='center')
                    elif cell.column == 8:
                        cell.alignment = Alignment(vertical='center', horizontal='right')
                    else:
                        cell.alignment = Alignment(vertical='center', horizontal='left')

            # Set column widths for better fit, name and featured columns fit to text
            name_col_idx = 2
            featured_col_idx = 6
            max_name_len = max(len(str(ws.cell(row=row, column=name_col_idx).value)) for row in range(header_row+1, ws.max_row+1)) if ws.max_row > header_row else 10
            name_col_width = min(max(15, max_name_len + 2), 40)  # min 15, max 40
            # Set featured column width to match header text length
            featured_header = headers[featured_col_idx - 1]
            featured_col_width = max(8, len(featured_header) + 4)  # min 8, header length + 2
            col_widths = [7, name_col_width, 15, 18, 18, featured_col_width, 10, 12]
            for i, width in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            filename = f"furniture_list_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response = FileResponse(output, as_attachment=True, filename=filename)
            return response
        elif export_format == 'pdf':
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.pdfgen import canvas
            from reportlab.lib.units import cm
            from reportlab.platypus import Table, TableStyle, Paragraph
            from reportlab.lib import colors
            from reportlab.lib.styles import getSampleStyleSheet
            import datetime
            output = io.BytesIO()
            page_size = landscape(A4)
            p = canvas.Canvas(output, pagesize=page_size)
            width, height = page_size
            styles = getSampleStyleSheet()
            # Title
            p.setFont('Helvetica-Bold', 18)
            p.setFillColor(colors.HexColor('#2d3748'))
            p.drawString(2 * cm, height - 2 * cm, 'Furniture List Export')
            # Timestamp
            p.setFont('Helvetica', 10)
            p.setFillColor(colors.HexColor('#4a5568'))
            timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            p.drawString(2 * cm, height - 2.7 * cm, f'Exported: {timestamp}')
            # Table Data
            headers = ['No.', 'Name', 'Item Code', 'Category', 'Brand', 'Featured', 'Visible', 'Price']
            data = [headers]
            for idx, item in enumerate(furniture_items, start=1):
                data.append([
                    str(idx),
                    item.name,
                    item.item_code,
                    item.category.name if item.category else '',
                    item.brand.name if item.brand else '',
                    'Yes' if item.featured else 'No',
                    'Yes' if item.is_visible else 'No',
                    str(item.price)
                ])
            # Table Style
            table_style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f6ad55')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),      # Name left
                ('ALIGN', (1, 0), (1, -1), 'LEFT'),      # Item Code left
                ('ALIGN', (2, 0), (5, -1), 'CENTER'),    # Category, Brand, Featured, Visible center
                ('ALIGN', (6, 0), (6, -1), 'RIGHT'),     # Price right
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('TOPPADDING', (0, 0), (-1, 0), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])
            # Dynamically set name column width based on max text length (tighter fit)
            max_name_len = max([len(str(row[1])) for row in data]) if data else 10
            name_col_width = min(max(2*cm, max_name_len*0.18*cm), 7*cm)  # Reduce minimum and maximum, and scale
            # Dynamically set featured column width
            max_featured_len = max([len(str(row[5])) for row in data]) if data else 3
            featured_col_width = min(max(1.2*cm, max_featured_len*0.35*cm), 2.5*cm)  # min 1.2cm, max 2.5cm
            col_widths = [1.5*cm, name_col_width, 3*cm, 4*cm, 4*cm, featured_col_width, 2*cm, 3*cm]

            # Estimate rows per page (landscape A4)
            # Further reduce margins to fit more rows
            row_height = 0.85 * cm  # Slightly smaller per row
            top_margin = 2.5 * cm   # Reduce top margin
            bottom_margin = 0.9 * cm  # Reduce bottom margin for footer text
            line_height = 0.6 * cm     # Reduce space for horizontal line above footer
            available_height = height - (top_margin + bottom_margin + line_height)
            rows_per_page = int(available_height // row_height) + 3
            # Always include header row on each page
            data_rows = data[1:]
            total_rows = len(data_rows)
            pages = [data_rows[i:i+rows_per_page-1] for i in range(0, total_rows, rows_per_page-1)]

            for page_num, page_rows in enumerate(pages):
                # Table data for this page: header + chunk
                page_data = [headers] + page_rows
                # Table style: alternating row colors
                page_table_style = TableStyle(table_style.getCommands())
                for i in range(1, len(page_data)):
                    bg_color = colors.HexColor('#f7fafc') if i % 2 == 0 else colors.white
                    page_table_style.add('BACKGROUND', (0, i), (-1, i), bg_color)
                table = Table(page_data, colWidths=col_widths)
                table.setStyle(page_table_style)
                # Draw table
                table_width, table_height = table.wrapOn(p, width, height)
                p.saveState()
                table.drawOn(p, 2 * cm, height - 4 * cm - table_height)
                p.restoreState()
                # Footer: horizontal line, text and page number on same row
                p.setLineWidth(0.5)
                p.setStrokeColor(colors.HexColor('#e2e8f0'))
                line_y = 1.7 * cm  # Slightly above text
                p.line(2 * cm, line_y, width - 2 * cm, line_y)
                # Footer text center, page number right on same row
                footer_y = 1.3 * cm
                p.setFont('Helvetica-Oblique', 9)
                p.setFillColor(colors.HexColor('#718096'))
                footer_text = "Generated by INHOUSE Portal"
                p.drawCentredString(width / 2, footer_y, footer_text)
                # Page number right-aligned on same row
                p.setFont('Helvetica', 8)
                p.setFillColor(colors.HexColor('#4a5568'))
                page_num_text = f"Page {page_num + 1} of {len(pages)}"
                p.drawRightString(width - 2 * cm, footer_y, page_num_text)
                p.showPage()
            p.save()
            output.seek(0)
            timestamp_str = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"furniture_list_{timestamp_str}.pdf"
            response = FileResponse(output, as_attachment=True, filename=filename)
            return response

    paginator = Paginator(furniture_items, 10) # Show 10 items per page
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj, 
        'items': page_obj.object_list, 
        'categories': categories,
        'brands': brands,
        'search_query': search_query,
        'selected_category': category_id,
        'selected_brand': brand_id,
        'selected_featured': featured_filter,
        'selected_is_visible':visible_filter,
    }
    return render(request, 'staff/furniture_list.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_furniture_manage(request, pk=None):
    if pk:
        furniture = get_object_or_404(Furniture, pk=pk)
        action = "Edit"
    else:
        furniture = None
        action = "Add"

    if request.method == 'POST':
        form = FurnitureForm(request.POST, request.FILES, instance=furniture)
        if form.is_valid():
            furniture_item = form.save()
            messages.success(request, f"Furniture item '{furniture_item.name}' saved successfully.")
            return redirect('staff_furniture_list')
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = FurnitureForm(instance=furniture)

    context = {
        'form': form,
        'action': action,
        'furniture': furniture, 
    }
    return render(request, 'staff/furniture_form.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_furniture_delete(request, pk):
    furniture = get_object_or_404(Furniture, pk=pk)
    if request.method == 'POST':
        furniture.delete()
        messages.success(request, f"Furniture item '{furniture.name}' deleted successfully.")
        return redirect('staff_furniture_list')
    context = {
        'furniture': furniture
    }
    return render(request, 'staff/furniture_confirm_delete.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_customer_list(request):
    # customers = CustomUser.objects.filter(is_staff=False).order_by('username')
    customers = CustomUser.objects.filter().order_by('username')

    search_query = request.GET.get('q', '').strip()
    selected_member_level = request.GET.get('member_level', '')

    filters = []

    if search_query:
        filters.append(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query) | 
            Q(c_code__icontains=search_query) 
        )

    # Apply Member Level Filter
    if selected_member_level:
        filters.append(Q(member_level=selected_member_level)) 

    # Apply all collected filters
    if filters:
        customers = customers.filter(*filters).distinct() # Use distinct() to avoid duplicates

    # Pagination
    paginator = Paginator(customers, 15) # Show 15 customers per page
    page_number = request.GET.get('page')
    try:
        page_obj = paginator.page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.page(1)
    except EmptyPage:
        page_obj = paginator.page(paginator.num_pages)

    context = {
        'page_obj': page_obj,
        'customers': page_obj.object_list, # The customers for the current page
        'search_query': search_query,
        'selected_member_level': selected_member_level,
        'member_level_choices': CustomUser.member_level_choices, # Get choices directly from CustomUser model
    }
    return render(request, 'staff/customer_list.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_customer_detail(request, pk):
    customer = get_object_or_404(CustomUser, pk=pk)
    customer_orders = customer.orders.all().order_by('-created_at')[:10]

    if request.method == 'POST':
        form = CustomUserUpdateForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            messages.success(request, f"Customer '{customer.username}' updated successfully.")
            return redirect('staff_customer_detail', pk=customer.pk)
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = CustomUserUpdateForm(instance=customer)

    context = {
        'customer': customer,
        'form': form,
        'customer_orders': customer_orders,
    }
    return render(request, 'staff/customer_detail.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_customer_create(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f"Customer '{user.username}' created successfully.")
            return redirect('staff_customer_list')
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = CustomUserCreationForm()

    context = {
        'form': form,
    }
    return render(request, 'staff/customer_form.html', context)


@login_required
@user_passes_test(is_staff_check)
def staff_customer_delete(request, pk):
    customer = get_object_or_404(CustomUser, pk=pk)
    if request.method == 'POST':
        customer_name_for_log = customer.username # Store name before deletion
        customer.delete()   

        messages.success(request, f"Customer '{customer_name_for_log}' deleted successfully.")
        return redirect('staff_customer_list')  
    context = {
        'customer': customer,
    }
    return render(request, 'staff/customer_confirm_delete.html', context) 

@login_required
@user_passes_test(is_staff_check)
def staff_memberlevel_list(request):
    from .models import MemberLevel
    levels = MemberLevel.objects.all().order_by('min_price')
    return render(request, 'staff/memberlevel_list.html', {'levels': levels})
@login_required
@user_passes_test(is_staff_check)
def staff_memberlevel_edit(request, pk=None):
    from .models import MemberLevel
    if pk:
        level = get_object_or_404(MemberLevel, pk=pk)
        action = "Edit"
    else:
        level = None
        action = "Add"

    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=level)
        if form.is_valid():
            member_level = form.save(commit=False)
            member_level.save()
            messages.success(request, f"Member Level '{member_level.name}' saved successfully.")
            return redirect('staff_memberlevel_list')
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = CategoryForm(instance=level)

    context = {
        'form': form,
        'action': action,
        'level': level,
    }
    return render(request, 'staff/memberlevel_form.html', context)  
@login_required
@user_passes_test(is_staff_check)
def staff_memberlevel_delete(request, pk):
    from .models import MemberLevel
    level = get_object_or_404(MemberLevel, pk=pk)
    if request.method == 'POST':
        level_name_for_log = level.name # Store name before deletion
        level.delete() 
        messages.success(request, f"Member Level '{level_name_for_log}' deleted successfully.")
        return redirect('staff_memberlevel_list')
    context = {
        'level': level,
    }
    return render(request, 'staff/memberlevel_confirm_delete.html', context) # Create this template


@login_required
@user_passes_test(is_staff_check)
def staff_memberlevel_create(request):
    from .models import MemberLevel
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            member_level = form.save(commit=False)
            member_level.save()
            messages.success(request, f"Member Level '{member_level.name}' created successfully.")
            return redirect('staff_memberlevel_list')
        else:
            messages.error(request, "Please correct the errors in the form.")
    else:
        form = CategoryForm()

    context = {
        'form': form,
    }
    return render(request, 'staff/memberlevel_form.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_mbl_list(request):
    membership_logs = UserMembershipLog.objects.all().select_related('user').order_by('-changed_at')

    search_query = request.GET.get('q')
    user_id = request.GET.get('user')

    if search_query:
        membership_logs = membership_logs.filter(
            Q(user__username__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    if user_id:
        membership_logs = membership_logs.filter(user__id=user_id)

    users = CustomUser.objects.all().order_by('username')

    context = {
        'membership_logs': membership_logs,
        'users': users,
        'search_query': search_query,
        'selected_user': user_id,
    }
    return render(request, 'staff/member_level_list.html', context)
@login_required
@user_passes_test(is_staff_check)
def staff_mbl_detail(request, pk):
   
    membership_log = get_object_or_404(UserMembershipLog.objects.select_related('user'), pk=pk)

    context = {
        'membership_log': membership_log,
    }
    return render(request, 'staff/member_log_detail.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_category_list(request):
    all_categories = Category.objects.select_related('parent').order_by('parent__name', 'name')
    top_level_categories = [cat for cat in all_categories if cat.parent is None]
    parent_groups = {}
    for cat in all_categories:
        if cat.parent:
            parent_groups.setdefault(cat.parent, []).append(cat)

    form = CategoryForm()
    if request.method == 'POST':
        form = CategoryForm(request.POST)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            return redirect('staff_category_list')

    context = {
        'top_level_categories': top_level_categories,
        'parent_groups': parent_groups,
        'form': form,
    }
    return render(request, 'staff/category_list.html', context)

@login_required
@user_passes_test(is_staff_check)
def staff_category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            category = form.save(commit=False)
            category.save()
            # log_category_save(sender=Category, instance=category, created=False, user=request.user)
            return redirect('staff_category_list')
    else:
        form = CategoryForm(instance=category)
    
    context = {
        'form': form,
        'category': category,
    }
    return render(request, 'staff/category_edit.html', context) # Create this template

@login_required
@user_passes_test(is_staff_check)
def staff_category_delete(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == 'POST':
        category_name_for_log = category.name # Store name before deletion
        category.delete()
      
        return redirect('staff_category_list')
    context = {
        'category': category,
    }
    return render(request, 'staff/category_confirm_delete.html', context) # Create this template


@login_required
@user_passes_test(is_staff_check)
def exchange_rate_list(request):
    rates = ExchangeRate.objects.all().order_by('currency')
    return render(request, 'exchange_rate/list.html', {'rates': rates})

@login_required
@user_passes_test(is_staff_check)
def create_exchange_rate(request):
    if request.method == 'POST':
        currency = request.POST.get('currency')
        rate = request.POST.get('rate')
        

        try:
            rate = float(rate)
        except (ValueError, TypeError):
            messages.error(request, "Invalid rate value. Please enter a valid number.")
            return render(request, 'exchange_rate/create.html', {'request_post': request.POST})

        if ExchangeRate.objects.filter(currency__iexact=currency).exists(): # use iexact for case-insensitive check
            messages.error(request, f"Exchange rate for {currency.upper()} already exists. Please update it instead.")
            return render(request, 'exchange_rate/create.html', {'request_post': request.POST})
            
        if rate <= 0:
            messages.error(request, "Exchange rate must be a positive number.")
            return render(request, 'exchange_rate/create.html', {'request_post': request.POST})

        currency = currency.upper().strip()

        try:
            new_rate = ExchangeRate(
                currency=currency,
                rate=rate,
                created_by=request.user
            )
            new_rate.save()
            messages.success(request, f"Exchange rate for {currency} created successfully.")
            return redirect('exchange_rate_list')
        except Exception as e: # Catch any other unexpected errors during save
            messages.error(request, f"An error occurred while saving the exchange rate: {e}")
            return render(request, 'exchange_rate/create.html', {'request_post': request.POST})

    # For GET request or if form validation fails and re-rendering
    return render(request, 'exchange_rate/create.html', {'request_post': request.POST if request.method == 'POST' else {}})

@login_required
@user_passes_test(is_staff_check)
def update_exchange_rate(request, currency):
    rate_obj = get_object_or_404(ExchangeRate, currency=currency)

    if request.method == 'POST':
        new_rate = request.POST.get('rate')
        try:
            new_rate = float(new_rate)
        except (ValueError, TypeError):
            messages.error(request, "Invalid rate value.")
            return redirect('update_exchange_rate', currency=currency)

        if new_rate != float(rate_obj.rate):
            old_rate = rate_obj.rate
            rate_obj.rate = new_rate
            rate_obj._old_rate = old_rate
            rate_obj._changed_by = request.user
            rate_obj.save()

            messages.success(request, f"{currency} rate updated from {old_rate} to {new_rate}")
        else:
            messages.info(request, "No changes detected.")

        return redirect('exchange_rate_list')

    return render(request, 'exchange_rate/update.html', {'rate': rate_obj})

from .models import ExchangeRateLog
from django.core.paginator import Paginator

@login_required
@user_passes_test(is_staff_check)
def exchange_rate_logs(request):
    logs = ExchangeRateLog.objects.select_related('changed_by').order_by('-timestamp')
    paginator = Paginator(logs, 25)
    page = request.GET.get('page')
    logs_page = paginator.get_page(page)
    return render(request, 'exchange_rate/logs.html', {'logs': logs_page})

@login_required
@user_passes_test(is_staff_check)
def db_backup(request):
    import datetime
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_filename = f"db_backup_{timestamp}.json"
    with tempfile.NamedTemporaryFile(delete=False, suffix='.json', mode='w', encoding='utf-8') as tmp:
        result = subprocess.run([
            'python', 'manage.py', 'dumpdata',
            '--natural-primary', '--natural-foreign', '--indent', '2'
        ], stdout=tmp, stderr=subprocess.PIPE, text=True, encoding='utf-8')
        tmp.flush()
        if result.returncode != 0:
            messages.error(request, f"Backup failed: {result.stderr}")
            return redirect('staff_dashboard')
        if os.path.getsize(tmp.name) == 0:
            messages.error(request, "Backup file is empty. No data exported.")
            return redirect('staff_dashboard')
        response = FileResponse(open(tmp.name, 'rb'), as_attachment=True, filename=backup_filename)
    return response
@login_required
@user_passes_test(is_staff_check)
def db_restore(request):
    preview_objects = None
    error_message = None
    file_name = None

    # Step 1: Handle file upload and preview
    if request.method == 'POST' and request.FILES.get('backup_file'):
        backup_file = request.FILES['backup_file']
        temp_path = os.path.join(settings.BASE_DIR, 'temp_restore.json')
        with open(temp_path, 'wb+') as destination:
            for chunk in backup_file.chunks():
                destination.write(chunk)
        try:
            with open(temp_path, 'r', encoding='utf-8') as f:
                preview_data = json.load(f)
            preview_objects = preview_data if isinstance(preview_data, list) else []
            file_name = backup_file.name
        except Exception as e:
            error_message = f"Could not read backup file: {e}"
            os.remove(temp_path)
            return render(request, 'staff/db_restore.html', {'error_message': error_message})
        # Show preview and confirmation form
        request.session['restore_temp_path'] = temp_path
        request.session['restore_file_name'] = file_name
        return render(request, 'staff/db_restore.html', {
            'preview_objects': preview_objects,
            'file_name': file_name,
            'error_message': error_message,
            'show_confirm': True,
        })

    # Step 2: Handle confirmation or cancellation
    elif request.method == 'POST' and (request.POST.get('confirm_restore') or request.POST.get('cancel_restore')):
        temp_path = request.session.get('restore_temp_path')
        file_name = request.session.get('restore_file_name')
        if not temp_path or not os.path.exists(temp_path):
            messages.error(request, "Restore file not found. Please upload again.")
            return redirect('staff_db_restore')
        if request.POST.get('confirm_restore'):
            result = subprocess.run(['python', 'manage.py', 'loaddata', temp_path])
            os.remove(temp_path)
            request.session.pop('restore_temp_path', None)
            request.session.pop('restore_file_name', None)
            if result.returncode == 0:
                messages.success(request, "Database restored successfully.")
            else:
                messages.error(request, "Restore failed.")
            return redirect('staff_dashboard')
        elif request.POST.get('cancel_restore'):
            os.remove(temp_path)
            request.session.pop('restore_temp_path', None)
            request.session.pop('restore_file_name', None)
            messages.info(request, "Restore cancelled.")
            return redirect('staff_db_restore')

    # Step 3: Initial GET request (show upload form)
    return render(request, 'staff/db_restore.html')